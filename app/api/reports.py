"""Export Reports API — generates PDF, XLSX, CSV reports with real data."""

import csv
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.models.catalog import (
    Order, OrderItem, User, Product, Payment, Category, Brand,
    Permission, Role,
)
from app.security import require_permission

router = APIRouter(prefix='/reports', tags=['Reports'])


class ReportRequest(BaseModel):
    report_type: str  # sales, revenue, orders, products, inventory, customers, payments
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    format: str = 'csv'  # csv, xlsx, pdf


def _parse_date(date_str: Optional[str]):
    if not date_str:
        return None
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        # Strip timezone info for naive DB columns
        return dt.replace(tzinfo=None)
    except (ValueError, AttributeError):
        return None


def _get_date_range(preset: str = "this_month"):
    now = datetime.utcnow()
    today = now.replace(hour=0, minute=0, second=0, microsecond=0)
    ranges = {
        "today": (today, now),
        "yesterday": (today - timedelta(days=1), today),
        "this_week": (today - timedelta(days=today.weekday()), now),
        "this_month": (today.replace(day=1), now),
        "last_month": (
            (today.replace(day=1) - timedelta(days=1)).replace(day=1),
            today.replace(day=1),
        ),
        "this_year": (today.replace(month=1, day=1), now),
    }
    return ranges.get(preset, ranges["this_month"])


async def _fetch_sales_data(db: AsyncSession, date_from, date_to):
    """Fetch detailed order data for reports."""
    stmt = (
        select(Order, User)
        .join(User, Order.user_id == User.id, isouter=True)
        .where(Order.status != 'cancelled')
    )
    if date_from:
        stmt = stmt.where(Order.created_at >= date_from)
    if date_to:
        stmt = stmt.where(Order.created_at <= date_to)
    stmt = stmt.order_by(Order.created_at.desc())

    result = await db.execute(stmt)
    rows = result.all()

    data = []
    for row in rows:
        o = row.Order
        u = row.User
        payment_method = ''
        payment_status = ''
        if hasattr(o, 'payment') and o.payment:
            payment_method = getattr(o.payment, 'payment_method', '') or ''
            payment_status = getattr(o.payment, 'status', '') or ''
        discount = float(o.subtotal or 0) + float(o.shipping_fee or 0) + float(o.tax or 0) - float(o.total_amount or 0)
        if discount < 0:
            discount = 0
        data.append({
            'order_number': o.order_number or '#{}'.format(o.id),
            'date': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else '',
            'customer': '{} {}'.format(u.first_name or '', u.last_name or '').strip() if u else 'Guest',
            'email': u.email if u else '',
            'subtotal': float(o.subtotal or 0),
            'discount': discount,
            'shipping': float(o.shipping_fee or 0),
            'tax': float(o.tax or 0),
            'total': float(o.total_amount or 0),
            'payment_method': payment_method,
            'payment_status': payment_status,
            'status': o.status or '',
        })

    summary = {
        'total_orders': len(data),
        'total_revenue': sum(d['total'] for d in data),
        'total_discounts': sum(d['discount'] for d in data),
        'total_shipping': sum(d['shipping'] for d in data),
        'total_tax': sum(d['tax'] for d in data),
        'avg_order_value': sum(d['total'] for d in data) / len(data) if data else 0,
    }
    return data, summary


async def _fetch_products_data(db: AsyncSession):
    stmt = select(Product).order_by(Product.name)
    result = await db.execute(stmt)
    products = result.scalars().all()
    data = []
    for p in products:
        data.append({
            'name': p.name or '',
            'sku': p.sku or '',
            'price': float(p.price or 0),
            'discount_price': float(p.discount_price or 0) if p.discount_price else '',
            'stock': p.stock or 0,
            'status': p.status or '',
            'is_featured': 'Yes' if p.is_featured else 'No',
            'created_at': p.created_at.strftime('%Y-%m-%d') if p.created_at else '',
        })
    summary = {
        'total_products': len(data),
        'active_products': sum(1 for d in data if d['status'] == 'active'),
        'total_stock': sum(d['stock'] for d in data),
        'low_stock': sum(1 for d in data if d['stock'] < 5),
    }
    return data, summary


async def _fetch_customers_data(db: AsyncSession, date_from, date_to):
    stmt = (
        select(User)
        .outerjoin(Role, User.role_id == Role.id)
        .where(or_(Role.permissions == None, Role.permissions < 16))
    )
    if date_from:
        stmt = stmt.where(User.created_at >= date_from)
    if date_to:
        stmt = stmt.where(User.created_at <= date_to)
    stmt = stmt.order_by(User.created_at.desc())
    result = await db.execute(stmt)
    users = result.scalars().all()

    data = []
    for u in users:
        order_count = 0
        total_spent = 0.0
        try:
            oc_result = await db.execute(
                select(func.count(), func.coalesce(func.sum(Order.total_amount), 0))
                .where(Order.user_id == u.id, Order.status != 'cancelled')
            )
            row = oc_result.one()
            order_count = row[0]
            total_spent = float(row[1])
        except Exception:
            pass

        data.append({
            'username': u.username or '',
            'email': u.email or '',
            'name': f"{u.first_name or ''} {u.last_name or ''}".strip(),
            'phone': u.phone or '',
            'orders': order_count,
            'total_spent': total_spent,
            'is_active': 'Active' if u.is_active else 'Inactive',
            'joined': u.created_at.strftime('%Y-%m-%d') if u.created_at else '',
        })
    summary = {
        'total_customers': len(data),
        'active_customers': sum(1 for d in data if d['is_active'] == 'Active'),
        'total_revenue': sum(d['total_spent'] for d in data),
        'avg_spend': sum(d['total_spent'] for d in data) / len(data) if data else 0,
    }
    return data, summary


async def _fetch_payments_data(db: AsyncSession, date_from, date_to):
    stmt = (
        select(Payment, Order, User)
        .join(Order, Payment.order_id == Order.id)
        .join(User, Order.user_id == User.id, isouter=True)
    )
    if date_from:
        stmt = stmt.where(Payment.created_at >= date_from)
    if date_to:
        stmt = stmt.where(Payment.created_at <= date_to)
    stmt = stmt.order_by(Payment.created_at.desc())
    result = await db.execute(stmt)
    rows = result.all()

    data = []
    for row in rows:
        p, o, u = row.Payment, row.Order, row.User
        data.append({
            'order_number': o.order_number if o else '',
            'customer': f"{u.first_name or ''} {u.last_name or ''}".strip() if u else 'Guest',
            'amount': float(p.amount or 0),
            'method': p.payment_method or '',
            'status': p.status or '',
            'date': p.created_at.strftime('%Y-%m-%d %H:%M') if p.created_at else '',
        })
    summary = {
        'total_payments': len(data),
        'total_amount': sum(d['amount'] for d in data),
        'completed': sum(1 for d in data if d['status'] == 'completed'),
        'pending': sum(1 for d in data if d['status'] == 'pending'),
    }
    return data, summary


def _to_csv(data: list[dict], report_type: str, summary: dict, date_from, date_to):
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(["ASAH'S PRIMENEST — " + report_type.upper() + " REPORT"])
    writer.writerow(["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    if date_from:
        writer.writerow(["From", date_from.strftime('%Y-%m-%d')])
    if date_to:
        writer.writerow(["To", date_to.strftime('%Y-%m-%d')])
    writer.writerow([])

    # Summary
    writer.writerow(["SUMMARY"])
    for k, v in summary.items():
        label = k.replace('_', ' ').title()
        if isinstance(v, float):
            writer.writerow([label, f"GHS {v:,.2f}"])
        else:
            writer.writerow([label, v])
    writer.writerow([])

    # Data
    if data:
        headers = [h.replace('_', ' ').title() for h in data[0].keys()]
        writer.writerow(headers)
        for row in data:
            values = []
            for v in row.values():
                if isinstance(v, float):
                    values.append(f"{v:,.2f}")
                else:
                    values.append(str(v))
            writer.writerow(values)
    else:
        writer.writerow(["No data found for the selected criteria."])

    output.seek(0)
    return output.getvalue()


def _to_xlsx(data: list[dict], report_type: str, summary: dict, date_from, date_to):
    """Generate XLSX file using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="XLSX export not available. Install openpyxl.")

    wb = Workbook()
    ws = wb.active
    ws.title = report_type.title()

    # Styles
    header_font = Font(name='Calibri', bold=True, size=14, color='F2660F')
    subtitle_font = Font(name='Calibri', size=10, color='999999')
    summary_header_font = Font(name='Calibri', bold=True, size=11, color='121010')
    col_header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    col_header_fill = PatternFill(start_color='F2660F', end_color='F2660F', fill_type='solid')
    cell_font = Font(name='Calibri', size=10)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = "ASAH'S PRIMENEST — " + report_type.upper() + " REPORT"
    ws['A1'].font = header_font
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = subtitle_font
    if date_from:
        ws['A3'] = f"From: {date_from.strftime('%Y-%m-%d')}"
        ws['A3'].font = subtitle_font
    if date_to:
        ws['B3'] = f"To: {date_to.strftime('%Y-%m-%d')}"
        ws['B3'].font = subtitle_font

    # Summary section
    row = 5
    ws.cell(row=row, column=1, value="SUMMARY").font = summary_header_font
    row += 1
    for k, v in summary.items():
        label = k.replace('_', ' ').title()
        ws.cell(row=row, column=1, value=label).font = cell_font
        if isinstance(v, float):
            ws.cell(row=row, column=2, value=f"GHS {v:,.2f}").font = cell_font
        else:
            ws.cell(row=row, column=2, value=v).font = cell_font
        row += 1
    row += 1

    # Column headers
    if data:
        headers = [h.replace('_', ' ').title() for h in data[0].keys()]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = col_header_font
            cell.fill = col_header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows
        for data_row in data:
            row += 1
            for col_idx, val in enumerate(data_row.values(), 1):
                cell = ws.cell(row=row, column=col_idx, value=f"{val:,.2f}" if isinstance(val, float) else str(val))
                cell.font = cell_font
                cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            try:
                if hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                if cell.value and not isinstance(cell, type(None)):
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _to_pdf(data: list[dict], report_type: str, summary: dict, date_from, date_to):
    """Generate PDF report using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        raise HTTPException(status_code=500, detail="PDF export not available. Install reportlab.")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    brand_orange = colors.HexColor('#F2660F')
    dark_bg = colors.HexColor('#121010')

    elements = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  textColor=brand_orange, fontSize=18, spaceAfter=6)
    elements.append(Paragraph("ASAH'S PRIMENEST", title_style))
    elements.append(Paragraph(f"{report_type.upper()} REPORT", styles['Heading2']))

    # Meta info
    meta_style = ParagraphStyle('Meta', parent=styles['Normal'], textColor=colors.grey, fontSize=9)
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", meta_style))
    if date_from:
        elements.append(Paragraph(f"From: {date_from.strftime('%Y-%m-%d')}", meta_style))
    if date_to:
        elements.append(Paragraph(f"To: {date_to.strftime('%Y-%m-%d')}", meta_style))
    elements.append(Spacer(1, 12))

    # Summary
    elements.append(Paragraph("Summary", styles['Heading3']))
    summary_data = [["Metric", "Value"]]
    for k, v in summary.items():
        label = k.replace('_', ' ').title()
        if isinstance(v, float):
            summary_data.append([label, f"GHS {v:,.2f}"])
        else:
            summary_data.append([label, str(v)])

    summary_table = Table(summary_data, colWidths=[200, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), brand_orange),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    # Data table
    if data:
        elements.append(Paragraph("Detailed Report", styles['Heading3']))
        headers = [h.replace('_', ' ').title() for h in data[0].keys()]

        table_data = [headers]
        for row in data[:500]:  # Limit rows for PDF
            table_row = []
            for v in row.values():
                if isinstance(v, float):
                    table_row.append(f"{v:,.2f}")
                else:
                    table_row.append(str(v)[:40])
            table_data.append(table_row)

        num_cols = len(headers)
        col_width = min((landscape(A4)[0] - 4*cm) / num_cols, 120)
        data_table = Table(table_data, colWidths=[col_width]*num_cols, repeatRows=1)
        data_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_orange),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9F9F9')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(data_table)
    else:
        elements.append(Paragraph("No data found for the selected criteria.", styles['Normal']))

    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], textColor=colors.grey, fontSize=8, alignment=TA_CENTER)
    elements.append(Paragraph("ASAH'S PRIMENEST — Confidential Business Report", footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get('/export')
async def export_report(
    report_type: str = Query("sales"),
    date_range: str = Query("this_month"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    fmt: str = Query("csv"),
    current_user: User = Depends(require_permission(Permission.VIEW)),
    db: AsyncSession = Depends(get_db),
):
    """Export report in CSV, XLSX, or PDF format with real data from PostgreSQL."""

    # Determine date range
    if date_from and date_to:
        dt_from = _parse_date(date_from)
        dt_to = _parse_date(date_to)
    else:
        dt_from, dt_to = _get_date_range(date_range)

    # Fetch data based on report type
    data, summary = [], {}
    if report_type in ('sales', 'revenue', 'orders'):
        data, summary = await _fetch_sales_data(db, dt_from, dt_to)
    elif report_type == 'products':
        data, summary = await _fetch_products_data(db)
    elif report_type == 'customers':
        data, summary = await _fetch_customers_data(db, dt_from, dt_to)
    elif report_type == 'payments':
        data, summary = await _fetch_payments_data(db, dt_from, dt_to)
    elif report_type == 'inventory':
        data, summary = await _fetch_products_data(db)
    else:
        data, summary = await _fetch_sales_data(db, dt_from, dt_to)

    report_label = report_type.upper()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"primenest_{report_type}_{timestamp}"

    if fmt == 'csv':
        content = _to_csv(data, report_type, summary, dt_from, dt_to)
        return StreamingResponse(
            io.BytesIO(content.encode('utf-8-sig')),
            media_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}.csv"'}
        )
    elif fmt == 'xlsx':
        buffer = _to_xlsx(data, report_type, summary, dt_from, dt_to)
        return StreamingResponse(
            buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}.xlsx"'}
        )
    elif fmt == 'pdf':
        buffer = _to_pdf(data, report_type, summary, dt_from, dt_to)
        return StreamingResponse(
            buffer,
            media_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{filename}.pdf"'}
        )
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {fmt}")
